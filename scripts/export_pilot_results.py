"""Exports pilot_run_report.json into a readable Excel workbook for diagnostic review.

This is a DRAFT/diagnostic export, separate from Coding_Sheet_RESULTS.xlsx (which is reserved for
the full 150-question run — writing 4-question pilot data into that structured template would
leave nearly every complexity-tier row empty and misrepresent it as a real result). This export's
job is different: let a human quickly see, per (experiment, question), exactly what was generated,
what every metric scored, and where metrics are legitimately not-applicable (None/missing) versus
genuinely zero — the two read very differently and get flattened into the same JSON structure
otherwise.

Run with: PYTHONPATH=src python scripts/export_pilot_results.py [path/to/pilot_run_report.json]
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from statistics import mean

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REPORT_PATH = PROJECT_ROOT / "pilot_run_report.json"
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "Pilot_Results_Draft.xlsx"

# Column order for the Detail sheet — every metric compute_all_metrics can ever produce, so a
# metric that's simply not applicable to a given experiment (e.g. context_recall for Direct LLM)
# renders as a genuinely blank cell, not a coerced 0.
METRIC_COLUMNS = [
    "answer_relevance", "exact_match", "f1_score", "semantic_similarity",
    "context_recall", "context_precision", "hit_at_k", "mrr",
    "faithfulness", "hallucination_rate", "evidence_coverage", "citation_correctness",
    "numerical_accuracy", "calculation_accuracy", "multi_step_reasoning_score", "explanation_completeness",
    "latency_seconds", "token_usage", "cost_per_answer_usd",
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_detail_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Detail (per run)"
    headers = [
        "Exp. No.", "Question ID", "Status", "Complexity", "Calls", "Total Tokens",
        "Elapsed (s)", "Reference Answer", "Generated Answer", *METRIC_COLUMNS, "Error",
    ]
    ws.append(headers)
    _style_header(ws)

    for r in rows:
        metrics = r.get("metrics", {})
        row_values = [
            r["exp_id"], r["question_id"], r["status"], r.get("complexity_used", ""),
            r.get("num_llm_calls", ""), r.get("total_tokens", ""), r.get("elapsed_seconds", ""),
            r.get("reference_answer", ""), r.get("generated_answer", ""),
            *[metrics.get(k, None) for k in METRIC_COLUMNS],
            r.get("error", ""),
        ]
        ws.append(row_values)

    widths = [10, 24, 10, 11, 7, 12, 11, 30, 45] + [12] * len(METRIC_COLUMNS) + [40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_experiment_summary_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Experiment Summary")
    headers = [
        "Exp. No.", "Runs", "OK", "Failed",
        "Avg Answer Relevance", "Avg F1", "Avg Semantic Similarity", "Avg Numerical Accuracy",
        "Avg Context Recall (RAG only)", "Avg Context Precision (RAG only)",
        "Runs with Verification Computed", "Avg Faithfulness (where computed)",
        "Avg Latency (s)", "Avg Tokens/Run", "Total Tokens",
    ]
    ws.append(headers)
    _style_header(ws)

    exp_ids = sorted({r["exp_id"] for r in rows}, key=lambda e: int(e.split("-")[1]))
    for exp_id in exp_ids:
        exp_rows = [r for r in rows if r["exp_id"] == exp_id]
        ok_rows = [r for r in exp_rows if r["status"] == "OK"]
        failed = len(exp_rows) - len(ok_rows)

        def avg_metric(key: str) -> float | None:
            values = [r["metrics"][key] for r in ok_rows if key in r.get("metrics", {})]
            return round(mean(values), 4) if values else None

        faithfulness_vals = [r["metrics"]["faithfulness"] for r in ok_rows if "faithfulness" in r.get("metrics", {})]
        tokens = [r.get("total_tokens", 0) for r in ok_rows]

        ws.append([
            exp_id, len(exp_rows), len(ok_rows), failed,
            avg_metric("answer_relevance"), avg_metric("f1_score"), avg_metric("semantic_similarity"),
            avg_metric("numerical_accuracy"), avg_metric("context_recall"), avg_metric("context_precision"),
            len(faithfulness_vals), round(mean(faithfulness_vals), 4) if faithfulness_vals else None,
            avg_metric("latency_seconds"), round(mean(tokens), 1) if tokens else None, sum(tokens),
        ])

    for i, w in enumerate([10, 7, 6, 7, 12, 10, 14, 14, 14, 15, 16, 16, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_metric_coverage_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    """Answers directly: is a metric legitimately not-applicable, or is it computed-but-zero?"""
    ws = wb.create_sheet("Metric Coverage Diagnostic")
    headers = ["Metric", "Runs Computed", "Runs N/A (not applicable)", "Zero Among Computed", "Sample Values"]
    ws.append(headers)
    _style_header(ws)

    ok_rows = [r for r in rows if r["status"] == "OK"]
    for key in METRIC_COLUMNS:
        values = [r["metrics"][key] for r in ok_rows if key in r.get("metrics", {})]
        zeros = sum(1 for v in values if v == 0)
        na = len(ok_rows) - len(values)
        sample = ", ".join(str(v) for v in values[:5])
        ws.append([key, len(values), na, zeros, sample])

    for i, w in enumerate([28, 15, 22, 18, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main(report_path: Path = DEFAULT_REPORT_PATH, output_path: Path = DEFAULT_OUTPUT_PATH) -> None:
    rows = json.loads(report_path.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    build_detail_sheet(wb, rows)
    build_experiment_summary_sheet(wb, rows)
    build_metric_coverage_sheet(wb, rows)
    wb.save(output_path)

    print(f"Written: {output_path}")
    print(f"  Detail: {len(rows)} rows")
    print(f"  Experiment Summary: {len({r['exp_id'] for r in rows})} experiments")


if __name__ == "__main__":
    report_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_REPORT_PATH
    main(report_arg)
