"""Generates `Groq_API_Call_Budget.xlsx`: per-experiment Groq call/token/cost estimates for a full
150-question FinanceBench run across all 14 experiments.

Call-count formulas per experiment are taken directly from `experiments/registry.py` and were
validated against a mocked GroqClient (see git history) — they are exact, not estimated. Token
counts per call are estimated from this codebase's actual system prompts (measured with tiktoken)
plus the real FinanceBench dataset's average question/answer length — see the "Methodology &
Assumptions" sheet in the generated workbook for exact figures and the one genuinely unknown input
(the Simple/Moderate/Complex question-complexity mix, which is only knowable after running Query
Understanding live against the dataset — see finagent-experiments skill and VALIDATION.md).

Run with: PYTHONPATH=src python scripts/generate_call_budget.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from finagent.config import (
    GROQ_PRICE_PER_MILLION_INPUT_TOKENS_USD,
    GROQ_PRICE_PER_MILLION_OUTPUT_TOKENS_USD,
    RETRIEVAL_TOP_K,
)

NUM_QUESTIONS = 150

# ---------------------------------------------------------------------------
# Measured token counts (tiktoken cl100k_base, matching config.py's chunk-sizing proxy)
# ---------------------------------------------------------------------------
AVG_QUESTION_TOKENS = 35  # measured: financebench_open_source.jsonl average
EVIDENCE_CHUNK_TOKENS = 525  # 500-token chunk (config.CHUNK_SIZE_TOKENS) + ~25 formatting overhead
EVIDENCE_BLOCK_TOKENS = EVIDENCE_CHUNK_TOKENS * RETRIEVAL_TOP_K  # worst-case: filtering keeps all 5

SYSTEM_PROMPT_TOKENS = {
    "EXP-01": 51, "EXP-02": 94, "EXP-03": 209, "EXP-04": 128, "EXP-05": 133, "EXP-06": 145,
    "query_understanding": 622,
    "query_refinement_refine": 168,
    "query_refinement_multi": 180,
    "reasoning": 497,
    "verification": 282,
}

# Estimated output tokens per call type (JSON structure size / typical answer length).
OUTPUT_TOKENS = {
    "EXP-01": 60, "EXP-02": 90, "EXP-03": 60, "EXP-04": 180, "EXP-05": 150, "EXP-06": 120,
    "query_understanding": 90,
    "query_refinement_refine": 25,
    "query_refinement_multi": 90,
    "reasoning": 200,
    "verification": 80,
}


def _call_tokens(call_type: str, *, includes_evidence: bool = False) -> tuple[int, int]:
    input_tokens = SYSTEM_PROMPT_TOKENS[call_type] + AVG_QUESTION_TOKENS
    if includes_evidence:
        input_tokens += EVIDENCE_BLOCK_TOKENS
    return input_tokens, OUTPUT_TOKENS[call_type]


# ---------------------------------------------------------------------------
# Per-experiment call plans: list of (call_type, includes_evidence) per route.
# Direct LLM (EXP-01..06) and non-adaptive RAG (EXP-07..10) have one fixed call plan.
# Adaptive experiments (EXP-11..14) have a Simple/Moderate/Complex call plan each.
# ---------------------------------------------------------------------------

def direct_llm_plan(exp_id: str) -> list[tuple[str, bool]]:
    return [(exp_id, False)]


FIXED_RAG_PLANS: dict[str, list[tuple[str, bool]]] = {
    "EXP-07": [("reasoning", True)],
    "EXP-08": [("query_understanding", False), ("reasoning", True)],
    "EXP-09": [("query_refinement_refine", False), ("reasoning", True)],
    "EXP-10": [("query_refinement_multi", False), ("reasoning", True)],
}

# Adaptive plans: (simple_plan, moderate_plan_no_refine, moderate_plan_refine, complex_plan_no_refine, complex_plan_refine)
# "no_refine"/"refine" variants reflect whether QueryAnalysis.needs_refinement fired for that question.
def adaptive_plans(include_refinement: bool, include_filtering: bool, include_verification: bool):
    del include_filtering  # filtering never costs a Groq call — included in signature for clarity only
    base = [("query_understanding", False)]
    tail = [("reasoning", True)] + ([("verification", True)] if include_verification else [])

    if not include_refinement:
        # EXP-12: capability disabled — every route behaves like Simple's query prep.
        simple = base + tail
        moderate = base + tail
        complex_ = base + tail
        return {"Simple": simple, "Moderate": moderate, "Complex (min)": complex_, "Complex (max)": complex_}

    simple = base + tail
    moderate_no_refine = base + tail
    moderate_refine = base + [("query_refinement_refine", False)] + tail
    complex_no_refine = base + [("query_refinement_multi", False)] + tail
    complex_refine = base + [("query_refinement_refine", False), ("query_refinement_multi", False)] + tail
    return {
        "Simple": simple,
        "Moderate (no refine)": moderate_no_refine,
        "Moderate (refine)": moderate_refine,
        "Complex (no refine, always multi-query)": complex_no_refine,
        "Complex (refine + multi-query)": complex_refine,
    }


def plan_tokens(plan: list[tuple[str, bool]]) -> tuple[int, int]:
    total_in = total_out = 0
    for call_type, includes_evidence in plan:
        i, o = _call_tokens(call_type, includes_evidence=includes_evidence)
        total_in += i
        total_out += o
    return total_in, total_out


def cost_usd(input_tokens: int, output_tokens: int) -> float:
    return (
        input_tokens / 1_000_000 * GROQ_PRICE_PER_MILLION_INPUT_TOKENS_USD
        + output_tokens / 1_000_000 * GROQ_PRICE_PER_MILLION_OUTPUT_TOKENS_USD
    )


# Assumed question-complexity mix — the one genuinely unknown input, pending a live run of Query
# Understanding against the real 150-question set. Chosen to roughly match FinanceBench's own
# question_type distribution (1/3 "metrics-generated" ~ direct lookups) — see methodology sheet.
COMPLEXITY_MIX = {"Simple": 0.45, "Moderate": 0.30, "Complex": 0.25}
# Within Moderate/Complex, assumed fraction that trigger needs_refinement.
MODERATE_REFINE_FRACTION = 0.5
COMPLEX_REFINE_FRACTION = 0.6

DIRECT_LLM_IDS = [f"EXP-{i:02d}" for i in range(1, 7)]
DIRECT_LLM_NAMES = {
    "EXP-01": "Direct LLM with Zero-Shot Prompting",
    "EXP-02": "Direct LLM with Role-Based Financial Analyst Prompting",
    "EXP-03": "Direct LLM with Few-Shot Prompting",
    "EXP-04": "Direct LLM with Stepwise Financial Reasoning Prompting",
    "EXP-05": "Direct LLM with Self-Verification Prompting",
    "EXP-06": "Direct LLM with Structured Output Prompting",
}
FIXED_RAG_NAMES = {
    "EXP-07": "Naïve RAG using ChromaDB",
    "EXP-08": "Metadata-Aware Naïve RAG using ChromaDB",
    "EXP-09": "Query-Rewritten RAG using ChromaDB",
    "EXP-10": "Multi-Query RAG using ChromaDB",
}
ADAPTIVE_NAMES = {
    "EXP-11": "Adaptive Multi-Agent RAG using ChromaDB",
    "EXP-12": "Adaptive Multi-Agent RAG without Query Refinement",
    "EXP-13": "Adaptive Multi-Agent RAG without Evidence Filtering",
    "EXP-14": "Adaptive Multi-Agent RAG without Verification Agent",
}
ADAPTIVE_FLAGS = {
    "EXP-11": dict(include_refinement=True, include_filtering=True, include_verification=True),
    "EXP-12": dict(include_refinement=False, include_filtering=True, include_verification=True),
    "EXP-13": dict(include_refinement=True, include_filtering=False, include_verification=True),
    "EXP-14": dict(include_refinement=True, include_filtering=True, include_verification=False),
}


def build_rows() -> list[dict]:
    rows: list[dict] = []

    for exp_id in DIRECT_LLM_IDS:
        plan = direct_llm_plan(exp_id)
        calls = len(plan)
        in_tok, out_tok = plan_tokens(plan)
        rows.append({
            "Exp. No.": exp_id, "Experiment Name": DIRECT_LLM_NAMES[exp_id], "Category": "Direct LLM",
            "Calls/Question (min)": calls, "Calls/Question (expected)": calls, "Calls/Question (max)": calls,
            "Total Calls (150 Q)": calls * NUM_QUESTIONS,
            "Avg Input Tokens/Q": in_tok, "Avg Output Tokens/Q": out_tok,
            "Total Tokens (150 Q)": (in_tok + out_tok) * NUM_QUESTIONS,
            "Est. Cost USD (150 Q)": round(cost_usd(in_tok, out_tok) * NUM_QUESTIONS, 4),
        })

    for exp_id, plan in FIXED_RAG_PLANS.items():
        calls = len(plan)
        in_tok, out_tok = plan_tokens(plan)
        rows.append({
            "Exp. No.": exp_id, "Experiment Name": FIXED_RAG_NAMES[exp_id], "Category": "RAG (fixed strategy)",
            "Calls/Question (min)": calls, "Calls/Question (expected)": calls, "Calls/Question (max)": calls,
            "Total Calls (150 Q)": calls * NUM_QUESTIONS,
            "Avg Input Tokens/Q": in_tok, "Avg Output Tokens/Q": out_tok,
            "Total Tokens (150 Q)": (in_tok + out_tok) * NUM_QUESTIONS,
            "Est. Cost USD (150 Q)": round(cost_usd(in_tok, out_tok) * NUM_QUESTIONS, 4),
        })

    for exp_id, flags in ADAPTIVE_FLAGS.items():
        plans = adaptive_plans(**flags)
        simple_calls = len(plans["Simple"])
        min_calls = min(len(p) for p in plans.values())
        max_calls = max(len(p) for p in plans.values())

        if flags["include_refinement"]:
            moderate_expected = len(plans["Moderate (no refine)"]) * (1 - MODERATE_REFINE_FRACTION) + \
                len(plans["Moderate (refine)"]) * MODERATE_REFINE_FRACTION
            complex_expected = len(plans["Complex (no refine, always multi-query)"]) * (1 - COMPLEX_REFINE_FRACTION) + \
                len(plans["Complex (refine + multi-query)"]) * COMPLEX_REFINE_FRACTION
        else:
            moderate_expected = len(plans["Moderate"])
            complex_expected = len(plans["Complex (min)"])

        blended_calls = (
            simple_calls * COMPLEXITY_MIX["Simple"]
            + moderate_expected * COMPLEXITY_MIX["Moderate"]
            + complex_expected * COMPLEXITY_MIX["Complex"]
        )

        # Token estimate: blend Simple/Moderate/Complex plans' token totals the same way.
        simple_tok = plan_tokens(plans["Simple"])
        if flags["include_refinement"]:
            moderate_tok_in, moderate_tok_out = plan_tokens(plans["Moderate (refine)"])
            complex_tok_in, complex_tok_out = plan_tokens(plans["Complex (refine + multi-query)"])
        else:
            moderate_tok_in, moderate_tok_out = plan_tokens(plans["Moderate"])
            complex_tok_in, complex_tok_out = plan_tokens(plans["Complex (min)"])

        blended_in = (
            simple_tok[0] * COMPLEXITY_MIX["Simple"]
            + moderate_tok_in * COMPLEXITY_MIX["Moderate"]
            + complex_tok_in * COMPLEXITY_MIX["Complex"]
        )
        blended_out = (
            simple_tok[1] * COMPLEXITY_MIX["Simple"]
            + moderate_tok_out * COMPLEXITY_MIX["Moderate"]
            + complex_tok_out * COMPLEXITY_MIX["Complex"]
        )

        rows.append({
            "Exp. No.": exp_id, "Experiment Name": ADAPTIVE_NAMES[exp_id],
            "Category": "Proposed / Ablation" if exp_id != "EXP-11" else "Proposed Full System",
            "Calls/Question (min)": min_calls,
            "Calls/Question (expected)": round(blended_calls, 2),
            "Calls/Question (max)": max_calls,
            "Total Calls (150 Q)": round(blended_calls * NUM_QUESTIONS),
            "Avg Input Tokens/Q": round(blended_in),
            "Avg Output Tokens/Q": round(blended_out),
            "Total Tokens (150 Q)": round((blended_in + blended_out) * NUM_QUESTIONS),
            "Est. Cost USD (150 Q)": round(cost_usd(blended_in, blended_out) * NUM_QUESTIONS, 4),
        })

    return rows


def write_workbook(rows: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Call Budget by Experiment"
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in rows:
        ws.append([row[h] for h in headers])
    for col_idx, header in enumerate(headers, start=1):
        width = max(len(header), 14) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    totals_row = ["TOTAL (all 14 experiments)", "", ""]
    totals_row += ["", "", ""]
    totals_row += [sum(r["Total Calls (150 Q)"] for r in rows)]
    totals_row += ["", ""]
    totals_row += [sum(r["Total Tokens (150 Q)"] for r in rows)]
    totals_row += [round(sum(r["Est. Cost USD (150 Q)"] for r in rows), 2)]
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("Methodology & Assumptions")
    notes = [
        ("Purpose", "Estimated Groq API call/token/cost budget for running all 14 experiments "
                    "against the full 150-question FinanceBench open-source split."),
        ("Call counts", "EXACT, not estimated, for the fixed-strategy experiments (EXP-01..10) — "
                         "taken directly from experiments/registry.py's PipelineConfig definitions "
                         "and validated against a mocked GroqClient. For adaptive experiments "
                         "(EXP-11..14), call counts per complexity route (Simple/Moderate/Complex) "
                         "are also exact; only the BLEND across routes depends on the assumed "
                         "question-complexity mix below."),
        ("Complexity mix (assumed)", f"Simple {COMPLEXITY_MIX['Simple']:.0%}, Moderate "
                                      f"{COMPLEXITY_MIX['Moderate']:.0%}, Complex {COMPLEXITY_MIX['Complex']:.0%} "
                                      "— the one genuinely unknown input. Chosen to roughly track "
                                      "FinanceBench's own question_type distribution (1/3 "
                                      "'metrics-generated', direct-lookup-leaning). Will be replaced "
                                      "with the REAL distribution once Query Understanding has "
                                      "actually run against all 150 questions during the pilot run."),
        ("Refinement trigger rate (assumed)", f"Moderate: {MODERATE_REFINE_FRACTION:.0%} of questions "
                                               f"trigger needs_refinement. Complex: {COMPLEX_REFINE_FRACTION:.0%}. "
                                               "Also assumption, not measured."),
        ("Token counts — system prompts", "Measured exactly via tiktoken (cl100k_base) against this "
                                           "codebase's actual prompt strings — see SYSTEM_PROMPT_TOKENS "
                                           "in generate_call_budget.py."),
        ("Token counts — question/answer", f"Average question length ({AVG_QUESTION_TOKENS} tokens) "
                                            "measured directly from financebench_open_source.jsonl's "
                                            "150 questions."),
        ("Token counts — evidence block", f"{EVIDENCE_CHUNK_TOKENS} tokens/chunk (config.CHUNK_SIZE_TOKENS "
                                           f"[500] + ~25 formatting overhead) × config.RETRIEVAL_TOP_K "
                                           f"[{RETRIEVAL_TOP_K}] = {EVIDENCE_BLOCK_TOKENS} tokens. This is "
                                           "a worst-case assumption (evidence filtering may return fewer "
                                           "than top_k chunks); real usage will likely run somewhat lower."),
        ("Token counts — output", "Estimated per call type based on expected JSON/answer structure "
                                   "size — not yet measured against a live model, since GROQ_API_KEY "
                                   "was not configured at the time this was generated. Re-run this "
                                   "script with measured values once live calls have been made."),
        ("Pricing", f"Groq llama-3.3-70b-versatile: ${GROQ_PRICE_PER_MILLION_INPUT_TOKENS_USD}/M input "
                     f"tokens, ${GROQ_PRICE_PER_MILLION_OUTPUT_TOKENS_USD}/M output tokens "
                     "(config.py, from groq.com/pricing — verify it hasn't changed before relying on "
                     "this for budgeting)."),
        ("Key efficiency finding", "Most metrics (Answer Quality, most of Financial Reasoning, all of "
                                    "Efficiency) cost ZERO additional Groq calls — they're computed from "
                                    "local embeddings and string/token comparison "
                                    "(finagent.metrics). Only the pipeline's own agent calls "
                                    "(Query Understanding, Query Refinement, Reasoning, Verification) "
                                    "consume budget; no metric requires a separate LLM-judge call."),
        ("Regenerate", "PYTHONPATH=src python scripts/generate_call_budget.py"),
    ]
    ws2.append(["Aspect", "Notes"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for aspect, note in notes:
        ws2.append([aspect, note])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 100
    for row in ws2.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[0].alignment = Alignment(vertical="top")

    wb.save(out_path)


if __name__ == "__main__":
    rows = build_rows()
    out_path = Path(__file__).resolve().parents[1] / "Groq_API_Call_Budget.xlsx"
    write_workbook(rows, out_path)
    print(f"Written: {out_path}")
    total_calls = sum(r["Total Calls (150 Q)"] for r in rows)
    total_cost = sum(r["Est. Cost USD (150 Q)"] for r in rows)
    print(f"Total Groq calls across all 14 experiments (150 Q each): {total_calls:,}")
    print(f"Total estimated cost: ${total_cost:.2f}")
