"""Writes experiment results into `Coding_Sheet_RESULTS.xlsx` (finagent-experiments skill §3/§4).

**Never writes to `Coding_Sheet.xlsx`** (the original reference workbook) — every public method
here operates on `config.RESULTS_WORKBOOK_COPY` exclusively, creating it as a copy of the original
on first use if it doesn't exist yet.

Sheet names are truncated to Excel's 31-character limit in the source workbook (one has a trailing
space); the exact strings are centralized in `_SHEET_*` constants here so a typo can't silently
create a new blank sheet instead of writing to the right one — see finagent-experiments skill §0.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from finagent.config import RESULTS_WORKBOOK_COPY, RESULTS_WORKBOOK_ORIGINAL
from finagent.data.schemas import QueryComplexity, QuestionResult

_SHEET_OVERALL_PERFORMANCE = "Overall Performance Results of "  # trailing space is intentional
_SHEET_RETRIEVAL_GROUNDING = "Retrieval and Evidence Groundin"
_SHEET_COMPARATIVE_RANKING = "Final Comparative Ranking and A"
_SHEET_COMPLEXITY_BREAKDOWN = "Query Complexity-Wise Final Res"

# Metric dict key (as produced by finagent.metrics.compute_all_metrics) -> workbook column header.
_OVERALL_METRIC_COLUMNS = {
    "answer_relevance": "Answer Relevance",
    "exact_match": "Exact Match",
    "f1_score": "F1-Score",
    "semantic_similarity": "Semantic Similarity",
    "numerical_accuracy": "Numerical Accuracy",
    "faithfulness": "Faithfulness",
    "hallucination_rate": "Hallucination Rate",
}
_RETRIEVAL_METRIC_COLUMNS = {
    "context_precision": "Context Precision",
    "context_recall": "Context Recall",
    "hit_at_k": "Hit@K",
    "mrr": "MRR",
    "evidence_coverage": "Evidence Coverage",
    "citation_correctness": "Citation Correctness",
}
_COMPLEXITY_METRIC_COLUMNS = {
    "answer_relevance": "Answer Relevance",
    "context_recall": "Context Recall",
    "numerical_accuracy": "Numerical Accuracy",
    "faithfulness": "Faithfulness",
    "hallucination_rate": "Hallucination Rate",
}


def _mean(values: list[float]) -> float | None:
    values = [v for v in values if v is not None]
    return sum(values) / len(values) if values else None


def _header_index_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    """Map column header text -> 1-indexed column number, from the sheet's header row."""
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value:
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def _find_row(ws: Worksheet, exp_col: int, exp_id: str, *, extra_col: int | None = None, extra_value: str | None = None) -> int | None:
    """Find the 1-indexed row matching `exp_id` (and optionally a second column's value)."""
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=exp_col).value or "").strip() != exp_id:
            continue
        if extra_col is not None:
            if str(ws.cell(row=row, column=extra_col).value or "").strip() != extra_value:
                continue
        return row
    return None


class ReportWriter:
    """Writes averaged per-experiment results into `Coding_Sheet_RESULTS.xlsx`'s four result sheets."""

    def __init__(self, path: str | Path = RESULTS_WORKBOOK_COPY) -> None:
        self._path = Path(path)
        self.ensure_workbook_exists()

    def ensure_workbook_exists(self) -> None:
        """Create `Coding_Sheet_RESULTS.xlsx` as a copy of the original if it doesn't exist yet.

        Raises:
            FileNotFoundError: if the original `Coding_Sheet.xlsx` is missing (nothing to copy).
        """
        if self._path.exists():
            return
        if not RESULTS_WORKBOOK_ORIGINAL.exists():
            raise FileNotFoundError(
                f"Cannot create results workbook: original not found at {RESULTS_WORKBOOK_ORIGINAL}"
            )
        shutil.copy2(RESULTS_WORKBOOK_ORIGINAL, self._path)

    def write_overall_performance(
        self,
        exp_id: str,
        question_results: list[QuestionResult],
        *,
        avg_latency_override: float | None = None,
        avg_token_usage_override: float | None = None,
    ) -> None:
        """Write averaged Answer-Quality/Grounding metrics into the `Overall Performance` sheet.

        Args:
            exp_id: Experiment ID, e.g. "EXP-07" — must match an existing row's `Exp. No.` value;
                this method updates that row's metric columns, it does not create new rows (the
                sheet is pre-populated with all 14 experiment rows already).
            question_results: All `QuestionResult`s for this experiment run.
            avg_latency_override: Precomputed average latency in seconds, if the caller wants to
                report it in a different unit/rounding than a raw mean of `metrics["latency_seconds"]`.
            avg_token_usage_override: Same, for token usage.
        """
        wb = openpyxl.load_workbook(self._path)
        ws = wb[_SHEET_OVERALL_PERFORMANCE]
        headers = _header_index_map(ws)
        row = _find_row(ws, headers["Exp. No."], exp_id)
        if row is None:
            raise ValueError(f"No row found for {exp_id!r} in sheet {_SHEET_OVERALL_PERFORMANCE!r}")

        for metric_key, column_name in _OVERALL_METRIC_COLUMNS.items():
            value = _mean([qr.metrics.get(metric_key) for qr in question_results])
            if value is not None and column_name in headers:
                ws.cell(row=row, column=headers[column_name], value=round(value, 4))

        avg_latency = avg_latency_override or _mean(
            [qr.metrics.get("latency_seconds") for qr in question_results]
        )
        if avg_latency is not None and "Avg. Latency" in headers:
            ws.cell(row=row, column=headers["Avg. Latency"], value=round(avg_latency, 3))

        avg_tokens = avg_token_usage_override or _mean(
            [qr.metrics.get("token_usage") for qr in question_results]
        )
        if avg_tokens is not None and "Avg. Token Usage" in headers:
            ws.cell(row=row, column=headers["Avg. Token Usage"], value=round(avg_tokens, 1))

        wb.save(self._path)

    def write_retrieval_grounding(
        self,
        exp_id: str,
        question_results: list[QuestionResult],
        *,
        top_k: int | None = None,
        wrong_company_year_cases: int | None = None,
        missing_evidence_cases: int | None = None,
    ) -> None:
        """Write averaged retrieval metrics into the `Retrieval and Evidence Grounding` sheet.

        Only applies to RAG experiments (EXP-07..14) — the sheet has no rows for EXP-01..06, and
        calling this for a Direct LLM experiment ID will raise since `_find_row` won't find one.

        Args:
            exp_id: Experiment ID, e.g. "EXP-07".
            question_results: All `QuestionResult`s for this experiment run.
            top_k: Retrieval depth used, if different from the sheet's pre-filled value (usually
                leave `None` — `Top-k Used` is typically populated once and left alone).
            wrong_company_year_cases: Count of questions where retrieval pulled evidence from the
                wrong company/year — tallied during error analysis, not computed automatically.
            missing_evidence_cases: Count of questions where no ground-truth evidence page was
                retrieved at all (`context_recall == 0`) — computed automatically if left `None`.
        """
        wb = openpyxl.load_workbook(self._path)
        ws = wb[_SHEET_RETRIEVAL_GROUNDING]
        headers = _header_index_map(ws)
        row = _find_row(ws, headers["Exp. No."], exp_id)
        if row is None:
            raise ValueError(f"No row found for {exp_id!r} in sheet {_SHEET_RETRIEVAL_GROUNDING!r}")

        for metric_key, column_name in _RETRIEVAL_METRIC_COLUMNS.items():
            value = _mean([qr.metrics.get(metric_key) for qr in question_results])
            if value is not None and column_name in headers:
                ws.cell(row=row, column=headers[column_name], value=round(value, 4))

        if top_k is not None and "Top-k Used" in headers:
            ws.cell(row=row, column=headers["Top-k Used"], value=top_k)

        if wrong_company_year_cases is not None and "Wrong Company/Year Retrieval Cases" in headers:
            ws.cell(row=row, column=headers["Wrong Company/Year Retrieval Cases"], value=wrong_company_year_cases)

        if missing_evidence_cases is None:
            missing_evidence_cases = sum(
                1 for qr in question_results if qr.metrics.get("context_recall") == 0.0
            )
        if "Missing Evidence Cases" in headers:
            ws.cell(row=row, column=headers["Missing Evidence Cases"], value=missing_evidence_cases)

        wb.save(self._path)

    def write_query_complexity_breakdown(self, exp_id: str, question_results: list[QuestionResult]) -> None:
        """Write per-complexity-tier averages into the `Query Complexity-Wise Final Results` sheet.

        The sheet is pre-populated with one row per (experiment, complexity tier) — this method
        updates the 3 rows matching `exp_id` (Simple/Moderate/Complex), never appends new rows.

        Args:
            exp_id: Experiment ID, e.g. "EXP-11".
            question_results: All `QuestionResult`s for this experiment run (mixed complexity
                tiers — this method groups them internally by `QuestionResult.complexity`).
        """
        wb = openpyxl.load_workbook(self._path)
        ws = wb[_SHEET_COMPLEXITY_BREAKDOWN]
        headers = _header_index_map(ws)

        by_complexity: dict[QueryComplexity, list[QuestionResult]] = {}
        for qr in question_results:
            by_complexity.setdefault(qr.complexity, []).append(qr)

        for complexity, group in by_complexity.items():
            row = _find_row(
                ws, headers["Exp. No."], exp_id,
                extra_col=headers["Query Complexity"], extra_value=complexity.value,
            )
            if row is None:
                raise ValueError(
                    f"No row found for {exp_id!r} / {complexity.value!r} in sheet {_SHEET_COMPLEXITY_BREAKDOWN!r}"
                )

            if "No. of Questions" in headers:
                ws.cell(row=row, column=headers["No. of Questions"], value=len(group))

            for metric_key, column_name in _COMPLEXITY_METRIC_COLUMNS.items():
                value = _mean([qr.metrics.get(metric_key) for qr in group])
                if value is not None and column_name in headers:
                    ws.cell(row=row, column=headers[column_name], value=round(value, 4))

            avg_latency = _mean([qr.metrics.get("latency_seconds") for qr in group])
            if avg_latency is not None and "Avg. Latency" in headers:
                ws.cell(row=row, column=headers["Avg. Latency"], value=round(avg_latency, 3))

        wb.save(self._path)

    def write_comparative_ranking_row(
        self,
        exp_id: str,
        *,
        answer_quality_score: float | None = None,
        retrieval_quality_score: float | None = None,
        grounding_score: float | None = None,
        financial_reasoning_score: float | None = None,
        efficiency_score: float | None = None,
        overall_weighted_score: float | None = None,
        improvement_over_naive_rag: float | None = None,
        improvement_over_direct_llm: float | None = None,
        final_rank: int | None = None,
        main_finding: str | None = None,
    ) -> None:
        """Write one row of the cross-experiment comparative synthesis sheet.

        Per finagent-experiments skill §3c, this sheet is filled LAST, only after every
        experiment's raw results (overall performance, retrieval grounding, complexity breakdown)
        are already recorded — the scores here are derived composites, not raw per-question
        averages, so the caller (a final synthesis pass, not a per-experiment run) is responsible
        for computing them from the other three sheets before calling this method.

        Args:
            exp_id: Experiment ID, e.g. "EXP-11".
            answer_quality_score..efficiency_score: Composite family scores (whatever weighting
                scheme the synthesis pass uses — this method just writes whatever is passed).
            overall_weighted_score: Final composite score across all 5 families.
            improvement_over_naive_rag: Relative delta vs. EXP-07.
            improvement_over_direct_llm: Relative delta vs. EXP-01.
            final_rank: 1-indexed rank among all 14 experiments.
            main_finding: Short free-text summary for this experiment's row.
        """
        wb = openpyxl.load_workbook(self._path)
        ws = wb[_SHEET_COMPARATIVE_RANKING]
        headers = _header_index_map(ws)
        row = _find_row(ws, headers["Exp. No."], exp_id)
        if row is None:
            raise ValueError(f"No row found for {exp_id!r} in sheet {_SHEET_COMPARATIVE_RANKING!r}")

        values = {
            "Answer Quality Score": answer_quality_score,
            "Retrieval Quality Score": retrieval_quality_score,
            "Grounding Score": grounding_score,
            "Financial Reasoning Score": financial_reasoning_score,
            "Efficiency Score": efficiency_score,
            "Overall Weighted Score": overall_weighted_score,
            "Improvement Over Naïve RAG": improvement_over_naive_rag,
            "Improvement Over Direct LLM": improvement_over_direct_llm,
            "Final Rank": final_rank,
            "Main Finding": main_finding,
        }
        for column_name, value in values.items():
            if value is not None and column_name in headers:
                ws.cell(row=row, column=headers[column_name], value=value)

        wb.save(self._path)
