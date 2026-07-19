"""Tests for report_writer.py — writes into a throwaway copy, never the project's real
Coding_Sheet_RESULTS.xlsx or Coding_Sheet.xlsx, so running this suite is always safe."""

from __future__ import annotations

import shutil

import openpyxl
import pytest

from finagent.config import RESULTS_WORKBOOK_ORIGINAL
from finagent.data.schemas import (
    Chunk,
    EvidenceItem,
    EvidenceReference,
    FinanceBenchQuestion,
    PipelineTrace,
    QueryAnalysis,
    QueryComplexity,
    QuestionResult,
    ReasoningOutput,
    VerificationResult,
)
from finagent.metrics import compute_all_metrics
from finagent.results.report_writer import ReportWriter

pytestmark = pytest.mark.skipif(
    not RESULTS_WORKBOOK_ORIGINAL.exists(), reason="Coding_Sheet.xlsx not present in this environment"
)


@pytest.fixture
def throwaway_workbook_path(tmp_path):
    """A tmp_path copy of the real Coding_Sheet.xlsx — every test writes only to this copy."""
    path = tmp_path / "Coding_Sheet_RESULTS_test.xlsx"
    shutil.copy2(RESULTS_WORKBOOK_ORIGINAL, path)
    return path


def _make_question_result(qid: str, complexity: QueryComplexity, exp_id: str) -> QuestionResult:
    question = FinanceBenchQuestion(
        question_id=qid, question="Q", reference_answer="$100 million",
        evidence=[EvidenceReference(doc_name="X_2022_10K", page_number=10, text="...")],
        company="X", document_type="10-K", document_name="X_2022_10K", document_year=2022,
        gics_sector="Tech", justification="", dataset_question_type="metrics-generated",
        assigned_complexity=complexity,
    )
    chunk = Chunk(chunk_id="C1", company="X", year=2022, report_type="10-K", section="Income Statement",
                  page_number=10, text="...", source_document="X_2022_10K.pdf")
    evidence = EvidenceItem(evidence_id="EV_001", chunk=chunk, relevance_score=0.8, retrieval_query="q")

    trace = PipelineTrace(experiment_id=exp_id, question=question, complexity_used=complexity)
    trace.retrieved_evidence = [evidence]
    trace.filtered_evidence = [evidence]
    trace.reasoning_output = ReasoningOutput(
        reasoning_steps=["s"], extracted_values={}, draft_answer="x", citations=["EV_001"],
        insufficient_evidence=False,
    )
    trace.verification_result = VerificationResult(passed=True, unsupported_claims=[], confidence=0.9, notes="")
    trace.generated_answer = "$100 million"
    trace.query_analysis = QueryAnalysis(
        complexity=complexity, company="X", year=2022, metric="m", question_type="lookup",
        needs_calculation=False, needs_multiple_evidence_chunks=False, needs_refinement=False,
        routing_rationale="",
    )
    trace.mark_finished()
    return QuestionResult(trace=trace, metrics=compute_all_metrics(trace, question))


class TestEnsureWorkbookExists:
    def test_creates_copy_if_missing(self, tmp_path):
        target = tmp_path / "new_results.xlsx"
        assert not target.exists()
        ReportWriter(path=target)
        assert target.exists()

    def test_does_not_overwrite_existing(self, throwaway_workbook_path):
        original_mtime = throwaway_workbook_path.stat().st_mtime
        ReportWriter(path=throwaway_workbook_path)
        assert throwaway_workbook_path.stat().st_mtime == original_mtime


class TestWriteOverallPerformance:
    def test_writes_correct_row_without_touching_others(self, throwaway_workbook_path):
        writer = ReportWriter(path=throwaway_workbook_path)
        results = [_make_question_result(f"q{i}", QueryComplexity.SIMPLE, "EXP-11") for i in range(3)]
        writer.write_overall_performance("EXP-11", results)

        wb = openpyxl.load_workbook(throwaway_workbook_path, data_only=True)
        ws = wb["Overall Performance Results of "]
        headers = {c.value: c.column for c in ws[1]}
        row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, headers["Exp. No."]).value == "EXP-11")

        assert ws.cell(row, headers["Answer Relevance"]).value is not None
        # A different experiment's row must remain untouched.
        other_row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, headers["Exp. No."]).value == "EXP-01")
        assert ws.cell(other_row, headers["Answer Relevance"]).value is None

    def test_unknown_exp_id_raises(self, throwaway_workbook_path):
        writer = ReportWriter(path=throwaway_workbook_path)
        with pytest.raises(ValueError):
            writer.write_overall_performance("EXP-99", [])


class TestWriteQueryComplexityBreakdown:
    def test_writes_all_three_complexity_rows(self, throwaway_workbook_path):
        writer = ReportWriter(path=throwaway_workbook_path)
        results = (
            [_make_question_result(f"s{i}", QueryComplexity.SIMPLE, "EXP-11") for i in range(3)]
            + [_make_question_result(f"m{i}", QueryComplexity.MODERATE, "EXP-11") for i in range(2)]
            + [_make_question_result(f"c{i}", QueryComplexity.COMPLEX, "EXP-11") for i in range(1)]
        )
        writer.write_query_complexity_breakdown("EXP-11", results)

        wb = openpyxl.load_workbook(throwaway_workbook_path, data_only=True)
        ws = wb["Query Complexity-Wise Final Res"]
        headers = {c.value: c.column for c in ws[1]}
        counts = {}
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, headers["Exp. No."]).value == "EXP-11":
                complexity = ws.cell(r, headers["Query Complexity"]).value
                counts[complexity] = ws.cell(r, headers["No. of Questions"]).value

        assert counts == {"Simple": 3, "Moderate": 2, "Complex": 1}


def test_original_workbook_never_modified(throwaway_workbook_path):
    """Sanity check: the writer only ever touches the copy passed to it, never the real original."""
    original_mtime_before = RESULTS_WORKBOOK_ORIGINAL.stat().st_mtime
    writer = ReportWriter(path=throwaway_workbook_path)
    results = [_make_question_result("q1", QueryComplexity.SIMPLE, "EXP-11")]
    writer.write_overall_performance("EXP-11", results)
    assert RESULTS_WORKBOOK_ORIGINAL.stat().st_mtime == original_mtime_before
